"""
Eval cases — CRUD + the four dataset sources:

    examples   1-click import of the agent's VERIFIED question→SQL pairs
    upload     CSV / XLSX / JSON files with flexible EN/FR headers
    manual     one form in the UI
    generated  generator.py proposals (start unverified — a human confirms)

The gold SQL is the ground truth of a case, so it deserves its own gate:
`verify_case` dry-runs it through the FULL validator+executor chain and only
then marks the case runnable. Mirrors evaluation/datasets/loader.py.
"""
from __future__ import annotations

import json
import re

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.models.data_eval import DataEvalCase
from app.services.evaluation.common import strip_accents

CATEGORIES = ["aggregation", "join", "filter", "date", "ranking",
              "grouping", "subquery", "insufficient"]


# ══════════════════════════════════════════════════════════════
#  CRUD
# ══════════════════════════════════════════════════════════════

def _case_dict(c: DataEvalCase) -> dict:
    return {
        "id": c.id, "question": c.question, "gold_sql": c.gold_sql,
        "expected_answer": c.expected_answer,
        "category": c.category or "filter",
        "difficulty": c.difficulty or "medium",
        "language": c.language,
        "source": c.source or "manual",
        "verified": bool(c.verified),
        "gold_note": c.gold_note,
        "created_at": str(c.created_at),
    }


def list_cases(db: Session, source_id: str) -> list:
    rows = (db.query(DataEvalCase)
            .filter(DataEvalCase.data_source_id == source_id)
            .order_by(DataEvalCase.created_at).all())
    return [_case_dict(c) for c in rows]


def _norm_case(raw: dict) -> dict | None:
    """Flexible key mapping so expert-written files 'just work'."""
    def pick(*keys):
        for k in keys:
            v = raw.get(k)
            if v not in (None, ""):
                return v
        return None
    q = pick("question", "q", "query_question")
    if not q or not str(q).strip():
        return None
    cat = str(pick("category") or "filter").lower()
    cat = cat if cat in CATEGORIES else "filter"
    sql = str(pick("gold_sql", "sql", "query", "gold") or "").strip() or None
    if not sql and cat != "insufficient":
        return None                       # gold SQL is the ground truth
    return {
        "question": str(q).strip(),
        "gold_sql": sql,
        "expected_answer": (str(pick("expected_answer", "ground_truth", "answer")
                                or "").strip() or None),
        "category": cat,
        "difficulty": str(pick("difficulty") or "medium").lower(),
        "language": (str(pick("language", "lang") or "").strip() or None),
    }


def _find_case(db: Session, source_id: str, case_id: str) -> DataEvalCase:
    c = (db.query(DataEvalCase)
         .filter(DataEvalCase.id == case_id,
                 DataEvalCase.data_source_id == source_id).first())
    if not c:
        raise HTTPException(404, "Test case not found")
    return c


def add_case(db: Session, source_id: str, data: dict,
             source: str = "manual", verified: bool = True) -> dict:
    norm = _norm_case(data)
    if not norm:
        raise HTTPException(400, "Question and gold SQL are required "
                                 "(SQL optional only for category=insufficient)")
    c = DataEvalCase(data_source_id=source_id, source=source,
                     verified=verified, **norm)
    db.add(c)
    db.commit()
    db.refresh(c)
    return _case_dict(c)


def update_case(db: Session, source_id: str, case_id: str, data: dict) -> dict:
    """Edit a case (typically fixing a generated gold SQL). Changing the SQL
    resets verification — the new ground truth must be re-proven."""
    c = _find_case(db, source_id, case_id)
    for k in ("question", "gold_sql", "expected_answer", "category",
              "difficulty", "language"):
        if k in data and data[k] is not None:
            setattr(c, k, data[k] or None)
    if "gold_sql" in data:
        c.verified = False
        c.gold_note = None
    db.commit()
    db.refresh(c)
    return _case_dict(c)


def delete_case(db: Session, source_id: str, case_id: str) -> dict:
    c = _find_case(db, source_id, case_id)
    db.delete(c)
    db.commit()
    return {"deleted": case_id}


def clear_cases(db: Session, source_id: str) -> dict:
    n = (db.query(DataEvalCase)
         .filter(DataEvalCase.data_source_id == source_id).delete())
    db.commit()
    return {"deleted": n}


def verify_case(db: Session, source_id: str, case_id: str) -> dict:
    """Dry-run the gold SQL through the full validator+executor chain.
    Success proves the ground truth; failure lands in gold_note."""
    from app.services.data_agent.nodes import run_validated
    c = _find_case(db, source_id, case_id)
    if not c.gold_sql:
        if c.category == "insufficient":
            c.verified = True
            c.gold_note = None
            db.commit()
            return _case_dict(c)
        raise HTTPException(400, "This case has no gold SQL")
    try:
        r = run_validated(source_id, c.gold_sql)
        c.verified = True
        c.gold_note = None
        note = f"{r['row_count']} row(s)"
    except Exception as e:
        c.verified = False
        c.gold_note = str(e)[:300]
        note = c.gold_note
    db.commit()
    db.refresh(c)
    d = _case_dict(c)
    d["verify_result"] = note
    return d


# ══════════════════════════════════════════════════════════════
#  Dataset sources
# ══════════════════════════════════════════════════════════════

def import_from_examples(db: Session, source_id: str) -> dict:
    """The agent's VERIFIED question→SQL pairs, copied as eval cases. They
    are already human-proven, so they arrive verified. Dedupes on the exact
    question text."""
    from app.models.data_agent import DataSourceExample
    existing = {c.question.strip().casefold() for c in
                db.query(DataEvalCase)
                .filter(DataEvalCase.data_source_id == source_id).all()}
    examples = (db.query(DataSourceExample)
                .filter(DataSourceExample.data_source_id == source_id,
                        DataSourceExample.is_verified.is_(True)).all())
    imported, skipped = [], 0
    for ex in examples:
        key = (ex.question or "").strip().casefold()
        if not key or key in existing:
            skipped += 1
            continue
        existing.add(key)
        c = DataEvalCase(data_source_id=source_id, question=ex.question.strip(),
                         gold_sql=ex.sql, source="examples", verified=True)
        db.add(c)
        imported.append(c)
    db.commit()
    return {"imported": len(imported), "skipped": skipped}


def _canon_header(h) -> str | None:
    """Spreadsheet header (EN/FR, friendly wording) → canonical key."""
    if h is None:
        return None
    k = re.sub(r"\s+", " ", strip_accents(str(h)).lower().replace("*", "").strip())
    aliases = {
        "question": "question", "questions": "question", "your question": "question",
        "sql": "gold_sql", "gold sql": "gold_sql", "query": "gold_sql",
        "requete": "gold_sql", "requete sql": "gold_sql",
        "expected sql": "gold_sql", "sql attendu": "gold_sql",
        "answer": "expected_answer", "expected answer": "expected_answer",
        "correct answer": "expected_answer", "reponse": "expected_answer",
        "reponse correcte": "expected_answer", "ground truth": "expected_answer",
        "category": "category", "categorie": "category",
        "type": "category", "question type": "category",
        "difficulty": "difficulty", "difficulte": "difficulty", "niveau": "difficulty",
        "language": "language", "langue": "language", "lang": "language",
    }
    return aliases.get(k)


def parse_dataset_file(filename: str, content: bytes) -> list:
    """.xlsx / .csv / .json → list of raw case dicts (loader.py machinery,
    gold-SQL headers)."""
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    def from_rows(header, rows):
        keys = [_canon_header(h) for h in header]
        out = []
        for row in rows:
            d = {}
            for k, v in zip(keys, row):
                if k and v not in (None, ""):
                    d[k] = v
            q = str(d.get("question") or "").strip()
            if not q or strip_accents(q.lower()).startswith(("example", "exemple")):
                continue
            out.append(d)
        return out

    if ext == "json":
        data = json.loads(content.decode("utf-8-sig"))
        if isinstance(data, dict):
            data = data.get("cases") or data.get("dataset") or []
        return data if isinstance(data, list) else []

    if ext == "csv":
        import csv
        import io
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        sample = text[:2000]
        delim = ";" if sample.count(";") > sample.count(",") else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delim))
        for i, r in enumerate(rows):
            if any(_canon_header(c) == "question" for c in r):
                return from_rows(r, rows[i + 1:])
        return []

    if ext in ("xlsx", "xlsm"):
        import io

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        for i, r in enumerate(rows):
            if any(_canon_header(c) == "question" for c in r):
                return from_rows(r, rows[i + 1:])
        return []

    raise HTTPException(400, f"Unsupported file type .{ext} — use .xlsx, .csv or .json")


def upload_dataset(db: Session, source_id: str, payload) -> dict:
    """A parsed list (or {"cases": [...]}) → cases, source='upload'.
    Uploaded gold SQL is trusted-but-unproven: cases arrive verified=False
    until a dry-run proves them (verify_case / the Verify button)."""
    if isinstance(payload, dict):
        payload = payload.get("cases") or payload.get("dataset") or []
    if not isinstance(payload, list):
        raise HTTPException(400, "Dataset must be a JSON list of test cases")
    imported, skipped = [], 0
    for raw in payload[:500]:
        norm = _norm_case(raw) if isinstance(raw, dict) else None
        if not norm:
            skipped += 1
            continue
        c = DataEvalCase(data_source_id=source_id, source="upload",
                         verified=False, **norm)
        db.add(c)
        imported.append(c)
    db.commit()
    return {"imported": len(imported), "skipped": skipped,
            "cases": [_case_dict(c) for c in imported]}


def dataset_template(db: Session, source) -> dict:
    """Raw JSON template listing the agent's ENABLED tables."""
    from app.models.data_agent import DataSourceTable
    tables = [f"{t.schema_name}.{t.table_name}" for t in
              db.query(DataSourceTable)
              .filter(DataSourceTable.data_source_id == source.id).all()
              if t.is_enabled]
    return {
        "_instructions": (
            "One object per test. 'question' is the natural-language question; "
            "'sql' is the CORRECT read-only SELECT that answers it (the ground "
            "truth). Use only the tables listed in _available_tables. "
            "category ∈ " + ", ".join(CATEGORIES) + " — 'insufficient' means "
            "the schema cannot answer and the agent should refuse (no sql)."
        ),
        "_available_tables": tables,
        "cases": [
            {"question": "How many orders were placed last month?",
             "sql": "SELECT COUNT(*) FROM orders WHERE created_at >= "
                    "date_trunc('month', now() - interval '1 month')",
             "category": "aggregation", "difficulty": "easy", "language": "en"},
        ],
    }
