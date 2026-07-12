"""
Table-element builders for CSV (pandas) and Excel (openpyxl + pandas).

Tabular data is NOT flattened to text — structure is preserved as a Table
Document Model with the same element envelope every parser uses:

    id, parent_id, type,
    hierarchy     {depth, level},
    location      {sheet, range},
    content       {table_name, columns[{name,type,nullable}], rows[{index,values}]},
    metadata      {order, confidence, row_count, column_count, token_estimate},
    relationships {parent, children, previous, next}

  CSV   → one `table` element                              (pandas)
  Excel → `workbook` → `sheet`* → `table` (+ `formula`,    (openpyxl + pandas)
          `merged_cell`) so citations keep sheet + cell range.

Row storage is capped (structure/citation view); the chunker still receives the
full data via the row/table sections built by the parsers.
"""

import logging

logger = logging.getLogger(__name__)

MAX_ELEMENT_ROWS = 1000     # cap rows embedded per table element
MAX_FORMULAS = 500          # cap formula elements per sheet
MAX_MERGED = 200            # cap merged-cell elements per sheet


def _tok(v) -> int:
    return max(1, len(str(v)) // 4)


def _col_type(dtype) -> str:
    s = str(dtype).lower()
    if "int" in s:
        return "integer"
    if "float" in s or "decimal" in s:
        return "number"
    if "bool" in s:
        return "boolean"
    if "datetime" in s or "date" in s or "time" in s:
        return "datetime"
    return "string"


def _columns(df):
    return [
        {"name": str(c), "type": _col_type(df[c].dtype), "nullable": bool(df[c].isna().any())}
        for c in df.columns
    ]


def _rows(df, max_rows=MAX_ELEMENT_ROWS):
    """JSON-safe rows [{index, values}]; NaN→null, dtypes normalized by to_json."""
    import json
    recs = json.loads(df.head(max_rows).to_json(orient="records", date_format="iso"))
    return [{"index": i + 1, "values": rec} for i, rec in enumerate(recs)]


def _table_text(name, columns, rows):
    col_names = [c["name"] for c in columns]
    lines = [f"Table: {name}"]
    for r in rows:
        v = r["values"]
        lines.append(" | ".join(f"{cn}: {v.get(cn, '')}" for cn in col_names))
    return "\n".join(lines)


def _table_element(eid, parent_id, table_name, df, depth, order, sheet=None, rng=None):
    columns = _columns(df)
    total = len(df)
    rows = _rows(df)
    meta = {"order": order, "confidence": 1.0, "row_count": total,
            "column_count": len(df.columns), "token_estimate": _tok(total * len(df.columns) * 4)}
    if total > len(rows):
        meta["rows_truncated"] = True
        meta["rows_shown"] = len(rows)
    return {
        "id": eid, "parent_id": parent_id, "type": "table",
        "hierarchy": {"depth": depth, "level": depth},
        "location": {"sheet": sheet, "range": rng},
        "content": {"table_name": table_name, "columns": columns, "rows": rows},
        "metadata": meta,
        "relationships": {"parent": parent_id, "children": [], "previous": None, "next": None},
    }, columns, rows


# ══════════════════════════════════════════════════════════════
#  CSV  (pandas)
# ══════════════════════════════════════════════════════════════

def build_csv_elements(df, table_name):
    """Return (elements, text_repr)."""
    el, columns, rows = _table_element("table_1", None, table_name, df, 1, 1)
    return [el], _table_text(table_name, columns, rows)


# ══════════════════════════════════════════════════════════════
#  Excel  (openpyxl + pandas)
# ══════════════════════════════════════════════════════════════

def _json_safe(v):
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v)


class _IdGen:
    def __init__(self):
        self.counters, self.n = {}, 0

    def nid(self, prefix):
        self.counters[prefix] = self.counters.get(prefix, 0) + 1
        return f"{prefix}_{self.counters[prefix]}"

    def order(self):
        self.n += 1
        return self.n


def _workbook_element(g, sheet_names, extra=None):
    content = {"sheet_count": len(sheet_names), "sheets": sheet_names,
               "semantic_text": f"Workbook with {len(sheet_names)} sheet(s): "
                                + ", ".join(sheet_names)}
    if extra:
        content.update(extra)
    return {
        "id": g.nid("workbook"), "parent_id": None, "type": "workbook",
        "hierarchy": {"depth": 1, "level": 1},
        "location": {"sheet": None, "range": None},
        "content": content,
        "metadata": {"order": g.order(), "confidence": 1.0, "token_estimate": _tok(sheet_names)},
        "relationships": {"parent": None, "children": [], "previous": None, "next": None},
    }


def _sheet_element(g, wb_id, name, index, rng, n_rows, n_cols):
    return {
        "id": g.nid("sheet"), "parent_id": wb_id, "type": "sheet",
        "hierarchy": {"depth": 2, "level": 2},
        "location": {"sheet": name, "range": rng},
        "content": {"name": name, "index": index,
                    "semantic_text": f"Sheet '{name}' with {n_rows} row(s) and {n_cols} column(s)"},
        "metadata": {"order": g.order(), "confidence": 1.0, "row_count": n_rows,
                     "column_count": n_cols, "token_estimate": 5},
        "relationships": {"parent": wb_id, "children": [], "previous": None, "next": None},
    }


def _cell_val(cells, r, c):
    cell = cells.get(f"{r},{c}")
    if cell is None:
        return None
    v = getattr(cell, "display_value", None)
    return v if v not in ("", None) else None


def _infer_type(vals):
    ne = [v for v in vals if v not in (None, "")]
    if not ne:
        return "string"

    def _isint(x):
        try:
            int(str(x).strip())
            return True
        except Exception:
            return False

    def _isfloat(x):
        try:
            float(str(x).strip())
            return True
        except Exception:
            return False

    if all(_isint(v) for v in ne):
        return "integer"
    if all(_isfloat(v) for v in ne):
        return "number"
    return "string"


def _coerce(v, t):
    if v is None:
        return None
    try:
        if t == "integer":
            return int(float(str(v)))
        if t == "number":
            return float(str(v))
    except Exception:
        pass
    return v


def _table_from_ks_structure(g, parent_id, sheet_name, ts, cells):
    """Build a typed table element from a ks-xlsx-parser detected table region."""
    cr = getattr(ts, "overall_range", None)
    try:
        tl, br = cr.top_left, cr.bottom_right
        r0, c0, r1, c1 = tl.row, tl.col, br.row, br.col
    except Exception:
        return None
    if r1 < r0 or c1 < c0:
        return None

    headers, seen = [], {}
    for j, c in enumerate(range(c0, c1 + 1)):
        v = _cell_val(cells, r0, c)
        h = str(v) if v is not None else f"col{j + 1}"
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        headers.append(h)

    col_vals = {h: [] for h in headers}
    rows = []
    for i, r in enumerate(range(r0 + 1, r1 + 1), 1):
        vals = {}
        for j, h in enumerate(headers):
            v = _cell_val(cells, r, c0 + j)
            vals[h] = v
            col_vals[h].append(v)
        rows.append({"index": i, "values": vals})

    columns = []
    for h in headers:
        t = _infer_type(col_vals[h])
        columns.append({"name": h, "type": t, "nullable": any(v is None for v in col_vals[h])})
        for row in rows:
            row["values"][h] = _coerce(row["values"].get(h), t)

    return {
        "id": g.nid("table"), "parent_id": parent_id, "type": "table",
        "hierarchy": {"depth": 3, "level": 3},
        "location": {"sheet": sheet_name, "range": str(cr)},
        "content": {"table_name": f"{sheet_name}!{cr}", "columns": columns, "rows": rows},
        "metadata": {"order": g.order(), "confidence": 1.0, "row_count": len(rows),
                     "column_count": len(columns), "token_estimate": _tok(len(rows) * len(columns) * 4)},
        "relationships": {"parent": parent_id, "children": [], "previous": None, "next": None},
    }


def _chart_element(g, parent_id, ch, sheet_name):
    series = []
    for sr in getattr(ch, "series", None) or []:
        series.append({
            "name": getattr(sr, "name", None),
            "name_ref": getattr(sr, "name_ref", None),
            "value_ref": getattr(sr, "value_ref", None) or getattr(sr, "val_ref", None),
        })
    ct = str(getattr(ch, "chart_type", "")).split(".")[-1].lower()
    title = getattr(ch, "title", None)
    summary = getattr(ch, "summary_text", None) or ""
    semantic = (f"{ct} chart" + (f" '{title}'" if title else "")
                + (f": {summary}" if summary else ""))
    anchor = getattr(ch, "anchor", None)
    return {
        "id": g.nid("chart"), "parent_id": parent_id, "type": "chart",
        "hierarchy": {"depth": 3, "level": 3},
        "location": {"sheet": sheet_name, "range": str(anchor) if anchor else None},
        "content": {"chart_type": ct, "title": title, "series": series,
                    "summary_text": summary, "semantic_text": semantic},
        "metadata": {"order": g.order(), "confidence": 1.0, "token_estimate": _tok(summary or semantic)},
        "relationships": {"parent": parent_id, "children": [], "previous": None, "next": None},
    }


def build_excel_workbook(file_path):
    """
    Return (elements, chunk_sections, engine). Primary: ks-xlsx-parser
    (RAG-native chunks + citations). Fallback: openpyxl + pandas.
    """
    try:
        return _build_excel_ks(file_path)
    except Exception as e:
        logger.warning(f"[EXCEL] ks-xlsx-parser unavailable/failed ({e}); using openpyxl")
        return _build_excel_openpyxl(file_path)


def _build_excel_ks(file_path):
    """Primary Excel extractor — ks-xlsx-parser (chunks, formulas, merges)."""
    import ks_xlsx_parser as kx
    import pandas as pd
    from collections import defaultdict
    from openpyxl.utils import get_column_letter

    res = kx.parse_workbook(file_path)
    sheets = res.workbook.sheets
    try:
        dfs = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
    except Exception:
        dfs = {}

    chunks_by_sheet = defaultdict(list)
    for ch in res.chunks:
        chunks_by_sheet[ch.sheet_name].append(ch)

    g = _IdGen()
    elements, chunk_sections = [], []
    sheet_names = [s.sheet_name for s in sheets]
    wb_el = _workbook_element(g, sheet_names, extra={
        "total_cells": getattr(res.workbook, "total_cells", None),
        "total_formulas": getattr(res.workbook, "total_formulas", None)})
    elements.append(wb_el)
    sheet_ids = []

    for s in sheets:
        name = s.sheet_name
        _ur = getattr(s, "used_range", None)
        rng = str(_ur) if _ur is not None else None
        df = dfs.get(name)
        n_rows = len(df) if df is not None else 0
        n_cols = len(df.columns) if df is not None else 0
        s_el = _sheet_element(g, wb_el["id"], name, getattr(s, "sheet_index", 0),
                              rng, n_rows, n_cols)
        elements.append(s_el)
        child_ids = []
        cells = getattr(s, "cells", None) or {}

        # ── Multiple tables per sheet: one element per detected table region.
        # Falls back to a single pandas grid for sheets ks didn't segment. ──
        sheet_ts = [t for t in (getattr(res.workbook, "table_structures", None) or [])
                    if getattr(t, "sheet_name", None) == name]
        made_table = False
        for ts in sheet_ts:
            t_el = _table_from_ks_structure(g, s_el["id"], name, ts, cells)
            if t_el:
                elements.append(t_el)
                child_ids.append(t_el["id"])
                made_table = True
        if not made_table and df is not None and not df.empty:
            tid = g.nid("table")
            t_el, cols, rows = _table_element(tid, s_el["id"], name, df, 3, g.order(),
                                              sheet=name, rng=rng)
            elements.append(t_el)
            child_ids.append(tid)

        # ── RAG chunks (the ks-xlsx-parser value-add) ──
        for ch in chunks_by_sheet.get(name, []):
            cid = g.nid("cell_block")
            text = ch.render_text or ""
            bt = str(getattr(ch, "block_type", "")).split(".")[-1].lower()
            tok = getattr(ch, "token_count", None)
            crange = str(ch.cell_range)
            elements.append({
                "id": cid, "parent_id": s_el["id"], "type": "cell_block",
                "hierarchy": {"depth": 3, "level": 3},
                "location": {"sheet": name, "range": crange},
                "content": {"block_type": bt, "cell_range": crange,
                            "text": text, "token_count": tok},
                "metadata": {"order": g.order(), "confidence": 1.0,
                             "token_estimate": tok or _tok(text)},
                "relationships": {"parent": s_el["id"], "children": [],
                                  "previous": getattr(ch, "prev_chunk_id", None),
                                  "next": getattr(ch, "next_chunk_id", None)},
            })
            child_ids.append(cid)
            chunk_sections.append((f"{name}!{ch.cell_range}", text))

        # formulas (from ks cells)
        n_f = 0
        for coord, cell in (getattr(s, "cells", None) or {}).items():
            f = getattr(cell, "formula", None)
            if not f:
                continue
            n_f += 1
            if n_f > MAX_FORMULAS:
                break
            a1 = getattr(cell, "a1_ref", None) or coord
            fid = g.nid("formula")
            elements.append({
                "id": fid, "parent_id": s_el["id"], "type": "formula",
                "hierarchy": {"depth": 3, "level": 3},
                "location": {"sheet": name, "range": a1},
                "content": {"cell": a1,
                            "formula": str(f) if str(f).startswith("=") else f"={f}",
                            "calculated_value": _json_safe(getattr(cell, "formula_value", None))},
                "metadata": {"order": g.order(), "confidence": 1.0, "token_estimate": _tok(f)},
                "relationships": {"parent": s_el["id"], "children": [], "previous": None, "next": None},
            })
            child_ids.append(fid)

        # merged regions
        for mr in (getattr(s, "merged_regions", None) or [])[:MAX_MERGED]:
            try:
                tl, br = mr.range.top_left, mr.range.bottom_right
                r = f"{get_column_letter(tl.col)}{tl.row}:{get_column_letter(br.col)}{br.row}"
                master = getattr(mr, "master", tl)
                mcell = (getattr(s, "cells", None) or {}).get(f"{master.row},{master.col}")
                val = _json_safe(getattr(mcell, "display_value", None)) if mcell else None
            except Exception:
                r, val = str(mr), None
            mid = g.nid("merged")
            elements.append({
                "id": mid, "parent_id": s_el["id"], "type": "merged_cell",
                "hierarchy": {"depth": 3, "level": 3},
                "location": {"sheet": name, "range": r},
                "content": {"range": r, "value": val},
                "metadata": {"order": g.order(), "confidence": 1.0, "token_estimate": 3},
                "relationships": {"parent": s_el["id"], "children": [], "previous": None, "next": None},
            })
            child_ids.append(mid)

        # charts on this sheet
        for ch in (getattr(res.workbook, "charts", None) or []):
            if getattr(ch, "sheet_name", None) != name:
                continue
            ce = _chart_element(g, s_el["id"], ch, name)
            elements.append(ce)
            child_ids.append(ce["id"])
            if ce["content"]["semantic_text"]:
                chunk_sections.append((f"{name} · chart", ce["content"]["semantic_text"]))

        s_el["relationships"]["children"] = child_ids
        sheet_ids.append(s_el["id"])

    wb_el["relationships"]["children"] = sheet_ids
    return elements, chunk_sections, "ks-xlsx-parser"


def _iter_formulas(ws, ws_v):
    out = []
    try:
        for row in ws.iter_rows():
            for cell in row:
                is_formula = getattr(cell, "data_type", None) == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("="))
                if not is_formula:
                    continue
                calc = None
                if ws_v is not None:
                    try:
                        calc = ws_v[cell.coordinate].value
                    except Exception:
                        calc = None
                out.append({"cell": cell.coordinate, "formula": str(cell.value),
                            "calculated_value": _json_safe(calc)})
                if len(out) >= MAX_FORMULAS:
                    return out
    except Exception:
        pass
    return out


def _build_excel_openpyxl(file_path):
    """Fallback Excel extractor — openpyxl (structure) + pandas (typed tables)."""
    from openpyxl import load_workbook
    import pandas as pd

    wb_f = load_workbook(file_path, data_only=False)
    try:
        wb_v = load_workbook(file_path, data_only=True)
    except Exception:
        wb_v = None
    dfs = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")

    g = _IdGen()
    elements, chunk_sections = [], []
    sheet_names = wb_f.sheetnames
    wb_el = _workbook_element(g, sheet_names)
    elements.append(wb_el)
    sheet_ids = []

    for idx, name in enumerate(sheet_names):
        ws = wb_f[name]
        ws_v = wb_v[name] if wb_v is not None and name in wb_v.sheetnames else None
        df = dfs.get(name)
        try:
            rng = ws.dimensions
        except Exception:
            rng = None
        n_rows = len(df) if df is not None else 0
        n_cols = len(df.columns) if df is not None else 0
        s_el = _sheet_element(g, wb_el["id"], name, idx, rng, n_rows, n_cols)
        elements.append(s_el)
        child_ids = []

        if df is not None and not df.empty:
            tid = g.nid("table")
            t_el, cols, rows = _table_element(tid, s_el["id"], name, df, 3, g.order(),
                                              sheet=name, rng=rng)
            elements.append(t_el)
            child_ids.append(tid)
            chunk_sections.append((name, _table_text(name, cols, rows)))

        for frm in _iter_formulas(ws, ws_v):
            fid = g.nid("formula")
            elements.append({
                "id": fid, "parent_id": s_el["id"], "type": "formula",
                "hierarchy": {"depth": 3, "level": 3},
                "location": {"sheet": name, "range": frm["cell"]},
                "content": frm,
                "metadata": {"order": g.order(), "confidence": 1.0, "token_estimate": _tok(frm["formula"])},
                "relationships": {"parent": s_el["id"], "children": [], "previous": None, "next": None},
            })
            child_ids.append(fid)

        merged = list(getattr(ws, "merged_cells", []).ranges) if hasattr(ws, "merged_cells") else []
        for rng_m in merged[:MAX_MERGED]:
            r = str(rng_m)
            try:
                val = _json_safe(ws[r.split(":")[0]].value)
            except Exception:
                val = None
            mid = g.nid("merged")
            elements.append({
                "id": mid, "parent_id": s_el["id"], "type": "merged_cell",
                "hierarchy": {"depth": 3, "level": 3},
                "location": {"sheet": name, "range": r},
                "content": {"range": r, "value": val},
                "metadata": {"order": g.order(), "confidence": 1.0, "token_estimate": 3},
                "relationships": {"parent": s_el["id"], "children": [], "previous": None, "next": None},
            })
            child_ids.append(mid)

        s_el["relationships"]["children"] = child_ids
        sheet_ids.append(s_el["id"])

    wb_el["relationships"]["children"] = sheet_ids
    return elements, chunk_sections, "openpyxl"
