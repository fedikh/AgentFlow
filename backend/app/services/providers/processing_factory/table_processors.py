"""
Table Processors — extraction de données tabulaires.

Formats : CSV, Excel (.xlsx/.xls)
Chaque processeur retourne : list[{type: "table", content: str (Markdown), page: int}]

Les tableaux sont convertis en Markdown pour que le LLM puisse les lire.
Les gros fichiers sont découpés en blocs de 20 lignes avec le header répété.
"""
import logging

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════
# CSV — pandas
# ══════════════════════════════════════════════════════

def extract_csv(file_path: str) -> list[dict]:
    """
    Lit un fichier CSV.
    Utilisation : employes.csv, transactions.csv, clients.csv

    Petit fichier (< 50 lignes) → un seul bloc Markdown.
    Gros fichier → blocs de 20 lignes, header répété en haut.

    Pourquoi répéter le header : sans le header, le LLM voit "25000"
    et ne sait pas si c'est un salaire, un code postal ou un montant.
    Avec le header "Salaire | 25000" → il comprend.
    """
    import pandas as pd

    # Essayer UTF-8 puis Latin-1
    try:
        df = pd.read_csv(file_path, encoding="utf-8")
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(file_path, encoding="latin-1")
        except Exception as e:
            logger.error(f"CSV read failed: {e}")
            return []

    if df.empty:
        return []

    # Nettoyer les noms de colonnes
    df.columns = [str(c).strip() for c in df.columns]

    # Supprimer les colonnes complètement vides
    df = df.dropna(axis=1, how="all")

    # Supprimer les lignes complètement vides
    df = df.dropna(how="all")

    content_blocks = []

    if len(df) <= 50:
        # Petit fichier → un seul bloc
        md = _dataframe_to_markdown(df)
        content_blocks.append({"type": "table", "content": md, "page": 1})
    else:
        # Gros fichier → blocs de 20 lignes
        chunk_size = 20
        for start in range(0, len(df), chunk_size):
            chunk_df = df.iloc[start:start + chunk_size]
            md = _dataframe_to_markdown(chunk_df)
            block_num = start // chunk_size + 1
            content_blocks.append({
                "type": "table",
                "content": f"[Rows {start + 1}-{start + len(chunk_df)} of {len(df)}]\n{md}",
                "page": block_num,
            })

    # Ajouter un résumé statistique pour les colonnes numériques
    numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
    if numeric_cols:
        summary_lines = ["[Data summary]"]
        summary_lines.append(f"Total rows: {len(df)}")
        summary_lines.append(f"Columns: {', '.join(df.columns.tolist())}")
        for col in numeric_cols[:5]:  # Max 5 colonnes numériques
            summary_lines.append(f"{col}: min={df[col].min()}, max={df[col].max()}, mean={df[col].mean():.2f}")
        content_blocks.append({
            "type": "text",
            "content": "\n".join(summary_lines),
            "page": 1,
        })

    return content_blocks


# ══════════════════════════════════════════════════════
# EXCEL — openpyxl (multi-feuilles)
# ══════════════════════════════════════════════════════

def extract_xlsx(file_path: str) -> list[dict]:
    """
    Lit un fichier Excel feuille par feuille.
    Utilisation : budget_2024.xlsx, grille_salaires.xlsx, pipeline_ventes.xlsx

    Chaque feuille est traitée séparément.
    Le nom de la feuille est ajouté comme contexte [Sheet: nom].
    Les formules sont résolues (data_only=True) — le RAG voit
    le résultat calculé, pas "=SUM(A1:A10)".
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.error(f"Excel read failed: {e}")
        return []

    content_blocks = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]

        # Lire toutes les lignes non-vides
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cell for cell in cells):
                rows.append(cells)

        if len(rows) < 2:
            continue

        headers = rows[0]
        data_rows = rows[1:]

        if len(data_rows) <= 50:
            # Petit sheet → un seul bloc
            md = _rows_list_to_markdown(headers, data_rows)
            content_blocks.append({
                "type": "table",
                "content": f"[Sheet: {sheet_name}]\n{md}",
                "page": sheet_idx,
            })
        else:
            # Gros sheet → blocs de 20 lignes
            chunk_size = 20
            for start in range(0, len(data_rows), chunk_size):
                chunk = data_rows[start:start + chunk_size]
                md = _rows_list_to_markdown(headers, chunk)
                content_blocks.append({
                    "type": "table",
                    "content": f"[Sheet: {sheet_name}, Rows {start + 1}-{start + len(chunk)} of {len(data_rows)}]\n{md}",
                    "page": sheet_idx,
                })

    return content_blocks


# ══════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════

def _dataframe_to_markdown(df) -> str:
    """Convertit un DataFrame pandas en Markdown."""
    headers = [str(h) for h in df.columns]
    md_lines = ["| " + " | ".join(headers) + " |"]
    md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for _, row in df.iterrows():
        cells = [str(v).strip() if v is not None else "" for v in row]
        md_lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(md_lines)


def _rows_list_to_markdown(headers: list[str], rows: list[list[str]]) -> str:
    """Convertit headers + rows en Markdown."""
    num_cols = len(headers)
    md_lines = ["| " + " | ".join(headers) + " |"]
    md_lines.append("| " + " | ".join(["---"] * num_cols) + " |")
    for row in rows:
        padded = row + [""] * (num_cols - len(row))
        md_lines.append("| " + " | ".join(padded[:num_cols]) + " |")
    return "\n".join(md_lines)
