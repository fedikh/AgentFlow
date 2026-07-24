"""
Evaluation — datasets: case CRUD, expert-friendly templates (interactive HTML
form, Excel with dropdowns, JSON), flexible file parsing (.xlsx/.csv/.json,
EN/FR headers), and dataset generation (Ragas testset generator with a
chunk-grounded LLM fallback).
"""
from __future__ import annotations

import json
import re

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.models.evaluation import EvalCase
from .common import (logger, CATEGORIES, PLAIN_TYPES, TYPE_BY_LABEL,
                     strip_accents, json_from, space_llm, docs_of)


# ══════════════════════════════════════════════════════════════
#  CRUD
# ══════════════════════════════════════════════════════════════

def _case_dict(c: EvalCase) -> dict:
    return {
        "id": c.id, "question": c.question,
        "expected_answer": c.expected_answer,
        "expected_document": c.expected_document,
        "expected_page": c.expected_page,
        "category": c.category or "semantic",
        "difficulty": c.difficulty or "medium",
        "language": c.language,
        "source": getattr(c, "source", None) or "manual",
        "created_at": str(c.created_at),
    }


def list_cases(db: Session, space_id: str) -> list:
    rows = (db.query(EvalCase).filter(EvalCase.rag_space_id == space_id)
            .order_by(EvalCase.created_at).all())
    return [_case_dict(c) for c in rows]


def _norm_case(raw: dict) -> dict | None:
    """Flexible key mapping so expert-written JSON 'just works'."""
    def pick(*keys):
        for k in keys:
            v = raw.get(k)
            if v not in (None, ""):
                return v
        return None
    q = pick("question", "q", "query")
    if not q or not str(q).strip():
        return None
    page = pick("expected_page", "page")
    try:
        page = int(page) if page is not None else None
    except (TypeError, ValueError):
        page = None
    docs = pick("expected_document", "expected_documents", "source", "document", "doc")
    if isinstance(docs, list):
        docs = docs[0] if docs else None
    cat = str(pick("category") or "semantic").lower()
    return {
        "question": str(q).strip(),
        "expected_answer": (str(pick("expected_answer", "ground_truth", "ground_truth_answer", "answer") or "").strip() or None),
        "expected_document": (str(docs or "").strip() or None),
        "expected_page": page,
        "category": cat if cat in CATEGORIES else "semantic",
        "difficulty": str(pick("difficulty") or "medium").lower(),
        "language": (str(pick("language", "lang") or "").strip() or None),
    }


def add_case(db: Session, space_id: str, data: dict, source: str = "manual") -> dict:
    norm = _norm_case(data)
    if not norm:
        raise HTTPException(400, "Question is required")
    c = EvalCase(rag_space_id=space_id, source=source, **norm)
    db.add(c)
    db.commit()
    db.refresh(c)
    return _case_dict(c)


def upload_dataset(db: Session, space_id: str, payload) -> dict:
    """Accept an expert-written dataset: a list of cases, or {"cases": [...]}."""
    if isinstance(payload, dict):
        payload = payload.get("cases") or payload.get("dataset") or []
    if not isinstance(payload, list):
        raise HTTPException(400, "Dataset must be a JSON list of test cases")
    imported, skipped = [], 0
    for raw in payload[:500]:
        norm = _norm_case(raw) if isinstance(raw, dict) else None
        if not norm:
            skipped += 1
            continue
        c = EvalCase(rag_space_id=space_id, source="upload", **norm)
        db.add(c)
        imported.append(c)
    db.commit()
    for c in imported:
        db.refresh(c)
    return {"imported": len(imported), "skipped": skipped,
            "cases": [_case_dict(c) for c in imported]}


def delete_case(db: Session, space_id: str, case_id: str) -> dict:
    c = db.query(EvalCase).filter(EvalCase.id == case_id,
                                  EvalCase.rag_space_id == space_id).first()
    if not c:
        raise HTTPException(404, "Test case not found")
    db.delete(c)
    db.commit()
    return {"deleted": case_id}


def clear_cases(db: Session, space_id: str) -> dict:
    n = db.query(EvalCase).filter(EvalCase.rag_space_id == space_id).delete()
    db.commit()
    return {"deleted": n}


# ══════════════════════════════════════════════════════════════
#  File parsing — whatever the expert sends back
# ══════════════════════════════════════════════════════════════

def _canon_header(h) -> str | None:
    """Map a spreadsheet header (EN/FR, friendly wording) to a canonical key."""
    if h is None:
        return None
    k = re.sub(r"\s+", " ", strip_accents(str(h)).lower().replace("*", "").strip())
    aliases = {
        "question": "question", "questions": "question", "your question": "question",
        "correct answer": "expected_answer", "answer": "expected_answer",
        "expected answer": "expected_answer", "ground truth": "expected_answer",
        "reponse correcte": "expected_answer", "reponse": "expected_answer",
        "bonne reponse": "expected_answer",
        "document containing the answer": "expected_document",
        "document": "expected_document", "source": "expected_document",
        "fichier": "expected_document", "expected document": "expected_document",
        "page": "expected_page",
        "question type": "category", "type": "category", "categorie": "category",
        "category": "category",
        "difficulty": "difficulty", "difficulte": "difficulty", "niveau": "difficulty",
        "language": "language", "langue": "language", "lang": "language",
    }
    return aliases.get(k)


def _plain_type_to_slug(v) -> str:
    if not v:
        return "semantic"
    s = str(v).strip().lower()
    if s in CATEGORIES:
        return s
    return TYPE_BY_LABEL.get(s, "semantic")


def parse_dataset_file(filename: str, content: bytes) -> list:
    """.xlsx / .csv / .json → list of raw case dicts. Friendly EN/FR headers,
    plain-language types, EXAMPLE rows skipped."""
    ext = (filename or "").rsplit(".", 1)[-1].lower()

    def from_rows(header, rows):
        keys = [_canon_header(h) for h in header]
        out = []
        for row in rows:
            d = {}
            for k, v in zip(keys, row):
                if k and v not in (None, ""):
                    d[k] = v
            q = str(d.get("question") or "").strip()
            if not q or strip_accents(q.lower()).startswith(("example", "exemple")):
                continue
            d["category"] = _plain_type_to_slug(d.get("category"))
            out.append(d)
        return out

    if ext == "json":
        data = json.loads(content.decode("utf-8-sig"))
        if isinstance(data, dict):
            data = data.get("cases") or data.get("dataset") or []
        for d in (data if isinstance(data, list) else []):
            if isinstance(d, dict) and d.get("category"):
                d["category"] = _plain_type_to_slug(d.get("category"))
        return data if isinstance(data, list) else []

    if ext == "csv":
        import csv, io
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        sample = text[:2000]
        delim = ";" if sample.count(";") > sample.count(",") else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delim))
        for i, r in enumerate(rows):
            if any(_canon_header(c) == "question" for c in r):
                return from_rows(r, rows[i + 1:])
        return []

    if ext in ("xlsx", "xlsm"):
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        for i, r in enumerate(rows):
            if any(_canon_header(c) == "question" for c in r):
                return from_rows(r, rows[i + 1:])
        return []

    raise HTTPException(400, f"Unsupported file type .{ext} — use .xlsx, .csv or .json")


# ══════════════════════════════════════════════════════════════
#  Templates for the domain expert
# ══════════════════════════════════════════════════════════════

def dataset_template(db: Session, space) -> dict:
    """Raw JSON template (for technical users / API integrations)."""
    docs = docs_of(db, space)
    example_doc = docs[0] if docs else "your-document.pdf"
    return {
        "_instructions": (
            "Fill one object per test question. 'question' is required. "
            "'expected_answer' is the correct answer (ground truth). "
            "'expected_document' must be one of the documents listed in "
            "_available_documents; 'expected_page' is optional. "
            "category ∈ " + ", ".join(CATEGORIES) + "."
        ),
        "_available_documents": docs,
        "cases": [
            {"question": "Who is assigned to SOFRECOM?",
             "expected_answer": "Eya Ben Fredj",
             "expected_document": example_doc,
             "expected_page": 2,
             "category": "entity_lookup",
             "difficulty": "medium",
             "language": "fr"},
        ],
    }


def template_excel(db: Session, space) -> bytes:
    """Excel template a domain expert can fill without technical knowledge:
    instructions, example rows, dropdowns with real documents + plain types."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    docs = docs_of(db, space)
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    lists = wb.create_sheet("Lists")
    for i, d in enumerate(docs, start=1):
        lists.cell(row=i, column=1, value=d)
    for i, (label, _slug) in enumerate(PLAIN_TYPES, start=1):
        lists.cell(row=i, column=2, value=label)
    lists.sheet_state = "hidden"

    title_font = Font(bold=True, size=14)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="111827")
    example_font = Font(italic=True, color="9CA3AF")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = f"Test questions — {getattr(space, 'name', '')}"
    ws["A1"].font = title_font
    ws.merge_cells("A2:G2")
    ws["A2"] = ("HOW TO FILL: one row per question. Write the question exactly like a real "
                "user would ask it, and the correct answer. If you know which document "
                "contains the answer, pick it from the dropdown. Delete the grey EXAMPLE "
                "rows before sending the file back. / COMMENT REMPLIR : une ligne par "
                "question, avec la bonne réponse. Supprimez les lignes EXEMPLE avant de renvoyer le fichier.")
    ws["A2"].alignment = wrap
    ws.row_dimensions[2].height = 46

    headers = ["Question *", "Correct answer *", "Document containing the answer",
               "Page", "Question type", "Difficulty", "Language"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = head_font
        c.fill = head_fill
    widths = [52, 42, 34, 8, 30, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    examples = [
        ["EXAMPLE — Who is assigned to SOFRECOM?", "Eya Ben Fredj",
         docs[0] if docs else "", 2, PLAIN_TYPES[2][0], "medium", "fr"],
        ["EXAMPLE — How many days of annual leave do employees get?",
         "21 days per year", docs[0] if docs else "", "", PLAIN_TYPES[0][0], "easy", "en"],
    ]
    for r, row in enumerate(examples, start=5):
        for cidx, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=cidx, value=v)
            c.font = example_font

    if docs:
        dv_doc = DataValidation(type="list", formula1=f"=Lists!$A$1:$A${len(docs)}",
                                allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv_doc)
        dv_doc.add("C5:C300")
    dv_type = DataValidation(type="list", formula1=f"=Lists!$B$1:$B${len(PLAIN_TYPES)}",
                             allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_type)
    dv_type.add("E5:E300")
    dv_diff = DataValidation(type="list", formula1='"easy,medium,hard"', allow_blank=True,
                             showDropDown=False)
    ws.add_data_validation(dv_diff)
    dv_diff.add("F5:F300")
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def expert_form_html(db: Session, space) -> str:
    """Self-contained HTML questionnaire — the expert opens it in a browser,
    fills a guided form (auto-saved locally), and downloads a perfectly
    formatted answer file to send back to IT."""
    docs = docs_of(db, space)
    name = getattr(space, "name", "RAG space")
    doc_opts = "".join(f"<option>{d}</option>" for d in docs)
    type_opts = "".join(f'<option value="{s}">{l}</option>' for l, s in PLAIN_TYPES)
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Questions — {name}</title>
<style>
 body{{font-family:system-ui,sans-serif;background:#f5f6f8;margin:0;color:#111827}}
 .wrap{{max-width:760px;margin:0 auto;padding:26px 16px 80px}}
 h1{{font-size:21px;margin:0 0 4px}} .sub{{color:#6b7280;font-size:13.5px;margin-bottom:18px;line-height:1.6}}
 .card{{background:#fff;border:1px solid #e5e7eb;border-radius:14px;padding:18px;margin-bottom:14px}}
 label{{display:block;font-size:12.5px;font-weight:700;margin:12px 0 4px}}
 label small{{font-weight:500;color:#6b7280}}
 input,select,textarea{{width:100%;box-sizing:border-box;border:1.5px solid #d1d5db;border-radius:9px;
   padding:9px 11px;font-size:14px;font-family:inherit}} textarea{{min-height:64px;resize:vertical}}
 .row{{display:flex;gap:10px}} .row>div{{flex:1}}
 button{{border:0;border-radius:10px;padding:11px 18px;font-size:14px;font-weight:700;cursor:pointer}}
 .add{{background:#111827;color:#fff;width:100%;margin-top:16px}}
 .dl{{background:#16a34a;color:#fff;width:100%;font-size:15px;padding:14px}}
 .q{{display:flex;gap:10px;align-items:flex-start;border:1px solid #e5e7eb;border-radius:10px;
   padding:10px 12px;margin-bottom:8px;background:#fff}}
 .q b{{font-size:13.5px}} .q .a{{color:#374151;font-size:12.5px;margin-top:2px}}
 .q .m{{color:#6b7280;font-size:11.5px;margin-top:2px}}
 .q button{{background:#fee2e2;color:#b91c1c;padding:4px 10px;font-size:12px}}
 .count{{font-weight:800}} .steps{{display:flex;gap:10px;margin-bottom:18px}}
 .step{{flex:1;background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:10px 12px;font-size:12px;color:#374151}}
 .step b{{display:block;font-size:12.5px;margin-bottom:2px}}
 .hint{{background:#eef2ff;color:#3730a3;border-radius:10px;padding:10px 12px;font-size:12.5px;margin-top:10px;display:none}}
</style></head><body><div class="wrap">
<h1>📝 Test questions — {name}</h1>
<div class="sub">Your IT team needs real questions from your domain to test the document assistant.
No technical knowledge needed — fill the form, then click the green button and send the
downloaded file back to IT.<br>
<i>Aucune connaissance technique nécessaire — remplissez le formulaire puis renvoyez le fichier téléchargé à l'équipe IT.</i></div>
<div class="steps">
 <div class="step"><b>1 · Write questions</b>Ask like a real user would</div>
 <div class="step"><b>2 · Give the correct answer</b>So the system can be graded</div>
 <div class="step"><b>3 · Download &amp; send back</b>One green button at the bottom</div>
</div>
<div class="card">
 <label>Your question * <small>— exactly as a user would ask it</small></label>
 <textarea id="q" placeholder="Example: Who is assigned to SOFRECOM?"></textarea>
 <label>The correct answer * <small>— short and precise</small></label>
 <textarea id="a" placeholder="Example: Eya Ben Fredj"></textarea>
 <label>Which document contains the answer? <small>— optional but very useful</small></label>
 <div class="row"><div><select id="d"><option value="">— I'm not sure —</option>{doc_opts}</select></div>
 <div style="max-width:110px"><input id="p" type="number" placeholder="Page"></div></div>
 <div class="row"><div><label>Question type</label><select id="t">{type_opts}</select></div>
 <div><label>Difficulty</label><select id="df"><option>easy</option><option selected>medium</option><option>hard</option></select></div>
 <div style="max-width:110px"><label>Language</label><select id="lg"><option>fr</option><option>en</option><option>ar</option><option>other</option></select></div></div>
 <button class="add" onclick="add()">➕ Add this question</button>
 <div class="hint" id="hint"></div>
</div>
<div class="card">
 <div style="font-size:14px;margin-bottom:10px">You added <span class="count" id="n">0</span> question(s).
 <span style="color:#6b7280;font-size:12.5px">Aim for at least 10 — the more, the better the test.</span></div>
 <div id="list"></div>
 <button class="dl" onclick="dl()">💾 Download my answers — send this file to IT</button>
</div>
<script>
const KEY="eval-{space.id}";let items=[];try{{items=JSON.parse(localStorage.getItem(KEY))||[]}}catch(e){{}}
function esc(s){{const d=document.createElement('div');d.textContent=s||'';return d.innerHTML}}
function save(){{localStorage.setItem(KEY,JSON.stringify(items))}}
function render(){{document.getElementById('n').textContent=items.length;
 document.getElementById('list').innerHTML=items.map((it,i)=>
 `<div class="q"><div style="flex:1"><b>${{esc(it.question)}}</b><div class="a">✔ ${{esc(it.expected_answer)}}</div>
 <div class="m">${{esc(it.expected_document||'document not specified')}}${{it.expected_page?' · p.'+it.expected_page:''}}</div></div>
 <button onclick="del(${{i}})">remove</button></div>`).join('')}}
function del(i){{items.splice(i,1);save();render()}}
function hint(msg,ok){{const h=document.getElementById('hint');h.style.display='block';
 h.textContent=msg;h.style.background=ok?'#dcfce7':'#fee2e2';h.style.color=ok?'#166534':'#991b1b';
 setTimeout(()=>h.style.display='none',3500)}}
function add(){{const q=document.getElementById('q').value.trim(),a=document.getElementById('a').value.trim();
 if(!q){{hint("Please write the question first.");return}}
 if(!a){{hint("Please write the correct answer — it's how the system gets graded.");return}}
 items.push({{question:q,expected_answer:a,
  expected_document:document.getElementById('d').value||null,
  expected_page:parseInt(document.getElementById('p').value)||null,
  category:document.getElementById('t').value,
  difficulty:document.getElementById('df').value,
  language:document.getElementById('lg').value}});
 save();render();document.getElementById('q').value='';document.getElementById('a').value='';
 document.getElementById('p').value='';hint("Question added ✓ — you can write the next one.",true)}}
function dl(){{if(!items.length){{hint("Add at least one question first.");return}}
 const blob=new Blob([JSON.stringify({{cases:items}},null,2)],{{type:'application/json'}});
 const a=document.createElement('a');a.href=URL.createObjectURL(blob);
 a.download='evaluation-questions-{name}.json'.replace(/\\s+/g,'-');a.click();
 hint("File downloaded ✓ — please send it back to your IT team.",true)}}
render();
</script></div></body></html>"""


# ══════════════════════════════════════════════════════════════
#  Generation — Ragas testset generator, chunk-LLM fallback
# ══════════════════════════════════════════════════════════════

def _generate_native(db, space, n: int) -> list:
    from sqlalchemy import text as T
    rows = db.execute(T("""
        SELECT c.content, c.page, d.file_name
        FROM chunks c JOIN documents d ON d.id = c.document_id
        WHERE c.rag_space_id = :sid AND length(c.content) > 200
        ORDER BY random() LIMIT :n
    """), {"sid": space.id, "n": max(3, min(n, 15))}).fetchall()
    if not rows:
        raise HTTPException(400, "No indexed chunks — process documents first")
    llm = space_llm(db, space, max_tokens=1600)
    created = []
    for r in rows[:n]:
        prompt = (
            "You create ONE evaluation test case for a RAG system from this "
            "document excerpt. Reply ONLY with JSON: {\"question\": str, "
            "\"expected_answer\": str (short, directly supported by the text), "
            f"\"category\": one of {CATEGORIES}, "
            "\"difficulty\": easy|medium|hard, \"language\": ISO code like fr/en}. "
            "Use the excerpt's language.\n\n"
            f"Document: {r.file_name} (page {r.page})\nExcerpt:\n{r.content[:1500]}"
        )
        try:
            out = json_from(getattr(llm.invoke(prompt), "content", ""))
            if not out or not out.get("question"):
                continue
            created.append(add_case(db, space.id, {
                **out,
                "expected_document": r.file_name,
                "expected_page": r.page,
            }, source="generated"))
        except Exception as e:
            logger.warning(f"[EVAL] native generation failed on one chunk: {e}")
    return created


def _generate_ragas(db, space, n: int) -> list | None:
    """Ragas TestsetGenerator over real chunks; None on any failure. 150s budget."""
    import concurrent.futures as cf

    def _work():
        from sqlalchemy import text as T
        from langchain_core.documents import Document
        from ragas.testset import TestsetGenerator
        from ragas.llms import LangchainLLMWrapper
        from ragas.embeddings import LangchainEmbeddingsWrapper
        from app.services.embedding_factory.resolver import resolve_embedding_config
        from app.services.embedding_factory.factory import get_embedder

        rows = db.execute(T("""
            SELECT c.content, c.page, d.file_name
            FROM chunks c JOIN documents d ON d.id = c.document_id
            WHERE c.rag_space_id = :sid AND length(c.content) > 300
            ORDER BY random() LIMIT 25
        """), {"sid": space.id}).fetchall()
        if len(rows) < 3:
            return None
        docs = [Document(page_content=r.content[:2500],
                         metadata={"file_name": r.file_name, "page": r.page})
                for r in rows]
        llm = LangchainLLMWrapper(space_llm(db, space, max_tokens=2500))
        conf = resolve_embedding_config(db, space)
        emb = LangchainEmbeddingsWrapper(get_embedder(
            conf["family"], conf["model"], conf.get("api_key", ""), conf.get("base_url", "")))
        gen = TestsetGenerator(llm=llm, embedding_model=emb)
        ds = gen.generate_with_langchain_docs(docs, testset_size=min(n, 10))
        out = []
        for s in ds.samples:
            d = s.eval_sample
            q = getattr(d, "user_input", None)
            if not q:
                continue
            out.append({
                "question": q,
                "expected_answer": getattr(d, "reference", None),
                "category": "semantic",
            })
        return out or None

    try:
        with cf.ThreadPoolExecutor(max_workers=1) as pool:
            raw = pool.submit(_work).result(timeout=150)
        if not raw:
            return None
        return [add_case(db, space.id, r, source="generated") for r in raw]
    except Exception as e:
        logger.warning(f"[EVAL] ragas testset generation unavailable ({e}) — native fallback")
        return None


def generate_cases(db: Session, space, n: int = 8) -> dict:
    cases = _generate_ragas(db, space, n)
    engine = "ragas"
    if not cases:
        cases = _generate_native(db, space, n)
        engine = "llm"
    if not cases:
        raise HTTPException(500, "Generation produced no cases (check the space LLM key)")
    return {"cases": cases, "engine": engine}
